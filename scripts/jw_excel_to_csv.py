#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════
 JW (藻岩山 THE JEWELS) Excel → CSV 変換スクリプト v2.0
═══════════════════════════════════════════════════════════════
 SVD-OS | Momentum Peaks
 
 ■ 確定列マッピング (全シート統一):
   Col  2: 日付
   Col  6: LUNCH 人数
   Col  7: LUNCH 料理売上 (税込)
   Col  9: LUNCH 飲料売上 (税込)
   Col 11: LUNCH 合計 (税込)
   Col 12: LUNCH 客単価
   Col 14: DINNER 人数
   Col 15: DINNER 料理売上 (税込)
   Col 17: DINNER 飲料売上 (税込)
   Col 19: DINNER 合計 (税込)
   Col 20: DINNER 客単価
   Col 27: T.O. 人数
   Col 28: T.O. 料理売上
   Col 30: T.O. 飲料売上
   Col 32: T.O. 合計
   Col 33: T.O. 客単価
   Col 43: 席料
   Col 44: 南京錠
   Col 45: 花束
   Col 39: カレー
   Col 47: 売上合計 (花束預り金除く)
 
 ■ データ行: Row 5〜（Row 2のCol2に日付がないor合計行は除外）
 ■ 宴会(bq)・ビアガーデン(bg): JWでは該当なし → 全て0
═══════════════════════════════════════════════════════════════
"""

import openpyxl
import os
import csv
import sys
from datetime import datetime

# ─── 確定列マッピング（絶対に変更しないこと） ───
COL_MAP = {
    'date':          2,
    'l_count':       6,
    'l_food':        7,
    'l_drink':       9,
    'l_total':      11,
    'l_avg':        12,
    'd_count':      14,
    'd_food':       15,
    'd_drink':      17,
    'd_total':      19,
    'd_avg':        20,
    'to_count':     27,
    'to_food':      28,
    'to_drink':     30,
    'to_total':     32,
    'to_avg':       33,
    'seat_fee':     43,
    'lock_fee':     44,
    'flower':       45,
    'morris_curry': 39,
    'grand_total':  47,
}

# CSV出力ヘッダー
CSV_HEADERS = [
    'date', 'weekday',
    'l_count', 'l_food', 'l_drink', 'l_total', 'l_avg',
    'd_count', 'd_food', 'd_drink', 'd_total', 'd_avg',
    'to_count', 'to_food', 'to_drink', 'to_total', 'to_avg',
    'bq_count', 'bq_food', 'bq_drink', 'bq_total', 'bq_avg',
    'bg_count', 'bg_food', 'bg_drink', 'bg_total', 'bg_avg',
    'seat_fee', 'lock_fee', 'flower', 'morris_curry',
    'grand_total'
]

WEEKDAY_JP = ['月', '火', '水', '木', '金', '土', '日']

EXCEL_DIR = '/Users/satoshiiga/dotfiles/SVD_L1_08_Restaurant_Sales/Mt.MOIWA'
OUTPUT_CSV = '/Users/satoshiiga/dotfiles/MomentumPeaks/data/JW_daily.csv'


def safe_int(value):
    """セル値を安全にintに変換"""
    if value is None:
        return 0
    if isinstance(value, str):
        value = value.strip()
        if value in ('', '-', '#DIV/0!', '#VALUE!', '#REF!', 'ランチ休業', 'None'):
            return 0
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0
    try:
        return int(round(float(value)))
    except (ValueError, TypeError):
        return 0


def parse_date(cell_value):
    """セル値を日付文字列 YYYY-MM-DD に変換"""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.strftime('%Y-%m-%d')
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
            try:
                return datetime.strptime(cell_value, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def get_weekday(date_str):
    """日付文字列から曜日を取得"""
    dt = datetime.strptime(date_str, '%Y-%m-%d')
    return WEEKDAY_JP[dt.weekday()]


def process_all_files():
    """全Excelファイルを処理してCSV出力"""
    all_rows = {}  # date -> row dict (重複排除用)
    
    # ファイル一覧取得
    excel_files = []
    for year_dir in sorted(os.listdir(EXCEL_DIR)):
        year_path = os.path.join(EXCEL_DIR, year_dir)
        if not os.path.isdir(year_path) or year_dir.startswith('.'):
            continue
        for f in sorted(os.listdir(year_path)):
            if f.endswith('.xlsx') and not f.startswith('~'):
                excel_files.append(os.path.join(year_path, f))
    
    print(f"📁 {len(excel_files)}ファイルを処理します")
    
    for fpath in excel_files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            print(f"  ❌ {fname}: {e}")
            continue
        
        for sname in wb.sheetnames:
            ws = wb[sname]
            sheet_count = 0
            
            for row_idx in range(5, ws.max_row + 1):
                date_str = parse_date(ws.cell(row=row_idx, column=COL_MAP['date']).value)
                if not date_str:
                    continue
                
                # 合計行のスキップ（日付が存在しないor明らかに合計）
                row_data = {'date': date_str, 'weekday': get_weekday(date_str)}
                
                # 確定列から全データ読み取り
                for key, col in COL_MAP.items():
                    if key == 'date':
                        continue
                    row_data[key] = safe_int(ws.cell(row=row_idx, column=col).value)
                
                # JWに宴会・BGは存在しない
                row_data['bq_count'] = 0
                row_data['bq_food'] = 0
                row_data['bq_drink'] = 0
                row_data['bq_total'] = 0
                row_data['bq_avg'] = 0
                row_data['bg_count'] = 0
                row_data['bg_food'] = 0
                row_data['bg_drink'] = 0
                row_data['bg_total'] = 0
                row_data['bg_avg'] = 0
                
                # grand_total が0の場合、各チャネルから再計算
                if row_data['grand_total'] == 0:
                    calc = (row_data['l_total'] + row_data['d_total'] + 
                            row_data['to_total'] + 
                            row_data['seat_fee'] + row_data['lock_fee'] + 
                            row_data['flower'] + row_data['morris_curry'])
                    if calc > 0:
                        row_data['grand_total'] = calc
                
                all_rows[date_str] = row_data
                sheet_count += 1
            
            print(f"  ✅ {fname} / {sname}: {sheet_count}日")
        
        wb.close()
    
    # 日付順にソートして出力
    sorted_dates = sorted(all_rows.keys())
    print(f"\n📊 合計 {len(sorted_dates)} 日のデータ")
    print(f"   期間: {sorted_dates[0]} 〜 {sorted_dates[-1]}")
    
    # CSV出力
    os.makedirs(os.path.dirname(OUTPUT_CSV), exist_ok=True)
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        for date_str in sorted_dates:
            writer.writerow(all_rows[date_str])
    
    print(f"✅ CSV出力完了: {OUTPUT_CSV}")
    
    # サマリー検証
    total_gt = sum(all_rows[d]['grand_total'] for d in sorted_dates)
    total_seat = sum(all_rows[d]['seat_fee'] for d in sorted_dates)
    total_lock = sum(all_rows[d]['lock_fee'] for d in sorted_dates)
    total_flower = sum(all_rows[d]['flower'] for d in sorted_dates)
    total_curry = sum(all_rows[d]['morris_curry'] for d in sorted_dates)
    total_to = sum(all_rows[d]['to_total'] for d in sorted_dates)
    total_l = sum(all_rows[d]['l_total'] for d in sorted_dates)
    total_d = sum(all_rows[d]['d_total'] for d in sorted_dates)
    
    print(f"\n=== データ検証 ===")
    print(f"LUNCH 売上合計:   ¥{total_l:>12,}")
    print(f"DINNER 売上合計:  ¥{total_d:>12,}")
    print(f"T.O. 売上合計:    ¥{total_to:>12,}")
    print(f"席料 合計:        ¥{total_seat:>12,}")
    print(f"南京錠 合計:      ¥{total_lock:>12,}")
    print(f"花束 合計:        ¥{total_flower:>12,}")
    print(f"カレー 合計:      ¥{total_curry:>12,}")
    print(f"GRAND TOTAL:      ¥{total_gt:>12,}")


if __name__ == '__main__':
    process_all_files()
